#处理excel文件
import os
from config.handle_filepath import data_dir_path
from openpyxl import load_workbook

class HandleExcel:
    """
    处理excel文件类
    """
    #处理文件必须传的值：file_name，sheet_name
    def __init__(self, file_name, sheet_name = None):
        self.file_name = os.path.join(data_dir_path, file_name)
        self.sheet_name = sheet_name

    #获取所有测试用例
    def get_all_case(self):
        #创建文件对象
        wb = load_workbook(self.file_name)

        #获取表单。默认第一个，其他表单需要传参
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]

        #获取第一列数据
        head_data = tuple(ws.iter_rows(max_row=1, values_only=True))[0]

        #获取测试数据
        body_data = tuple(ws.iter_rows(min_row=2, values_only=True))

        #拼接测试数据，[{第一行测试数据}，{第二行测试数据}]
        all_list = []
        for i in body_data:
            all_list.append(dict(zip(head_data, i)))

        return all_list


if __name__ == '__main__':
    do_excel = HandleExcel("case.xlsx")
    a = do_excel.get_all_case()
    print(a)
